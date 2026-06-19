"""
Dosya Adı: Lcd_Reader.py

LCD Dijital Ekran Okuyucu v5 (Sadece Özel Model)
========================================================
- EasyOCR tamamen kaldırıldı, sistem çok daha hızlıdır.
- Sadece eğitilmiş özel model (CNN) kullanılır.
- Modelin okuduğu ham rakamlar birleştirilir ve 10'a bölünerek
  (ondalıklı sayı olarak) Excel'e kaydedilir.

Tuşlar:
  [S]  -> Kaydet
  [A]  -> Otomatik kayıt aç/kapat
  [C]  -> Kalibrasyon
  [+]  -> Aralığı artır
  [-]  -> Aralığı azalt
  [Q]  -> Çıkış
"""

import cv2
import numpy as np
import openpyxl
import os
import time
import json

# ── PyTorch (özel model) ─────────────────────────
try:
    import torch
    import torch.nn as nn
    TORCH_OK = True
except ImportError:
    print("[HATA] PyTorch kurulu degil!")
    TORCH_OK = False

CONFIG_FILE    = "lcd_config.json"
EXCEL_FILE     = "lcd_kayitlar.xlsx"
CUSTOM_MODEL   = "digit_model.pth"
SKLEARN_MODEL  = "digit_model_sklearn.pkl"

# ================================================================
#  Özel CNN modeli
# ================================================================
class DigitCNN(nn.Module):
    def __init__(self, n_classes=10):
        super().__init__()
        self.features = nn.Sequential(
            nn.Conv2d(1, 16, 3, padding=1), nn.ReLU(), nn.MaxPool2d(2),
            nn.Conv2d(16, 32, 3, padding=1), nn.ReLU(), nn.MaxPool2d(2),
            nn.Conv2d(32, 64, 3, padding=1), nn.ReLU(), nn.MaxPool2d(2),
        )
        self.classifier = nn.Sequential(
            nn.Flatten(),
            nn.Linear(64 * 4 * 4, 128), nn.ReLU(), nn.Dropout(0.4),
            nn.Linear(128, n_classes),
        )

    def forward(self, x):
        return self.classifier(self.features(x))

# ================================================================
#  Model yükleyici
# ================================================================
class CustomDigitModel:
    def __init__(self):
        self.model   = None
        self.classes = None
        self.backend = None
        self._load()

    def _load(self):
        if TORCH_OK and os.path.exists(CUSTOM_MODEL):
            try:
                ckpt = torch.load(CUSTOM_MODEL, map_location="cpu", weights_only=False)
                self.classes = ckpt["classes"]
                net = DigitCNN(n_classes=ckpt["n_classes"])
                net.load_state_dict(ckpt["model_state"])
                net.eval()
                self.model   = net
                self.backend = "pytorch"
                print(f"[OK] Ozel model yuklendi (PyTorch) siniflar={self.classes}")
                return
            except Exception as e:
                print(f"[!] PyTorch model yuklenemedi: {e}")

        if os.path.exists(SKLEARN_MODEL):
            try:
                import pickle
                with open(SKLEARN_MODEL, "rb") as f:
                    data = pickle.load(f)
                self.model   = data["model"]
                self.classes = data["classes"]
                self.backend = "sklearn"
                print(f"[OK] Ozel model yuklendi (sklearn) siniflar={self.classes}")
                return
            except Exception as e:
                print(f"[!] Sklearn model yuklenemedi: {e}")

    def is_ready(self):
        return self.model is not None

    def predict(self, roi_gray_32x32):
        if not self.is_ready(): return None, 0.0
        img = roi_gray_32x32.astype(np.float32) / 255.0
        if self.backend == "pytorch":
            tensor = torch.tensor(img).unsqueeze(0).unsqueeze(0)
            with torch.no_grad():
                logits = self.model(tensor)
                probs  = torch.softmax(logits, dim=1)[0]
            best_idx   = probs.argmax().item()
            return self.classes[best_idx], probs[best_idx].item()
        elif self.backend == "sklearn":
            flat  = img.reshape(1, -1)
            idx   = self.model.predict(flat)[0]
            prob  = self.model.predict_proba(flat)[0]
            return self.classes[idx], prob.max()
        return None, 0.0

# ================================================================
#  Görüntü ön işleme
# ================================================================
def preprocess_digit(roi):
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY) if len(roi.shape) == 3 else roi
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(4, 4))
    gray = clahe.apply(gray)
    return cv2.resize(gray, (32, 32), interpolation=cv2.INTER_CUBIC)

custom_model = CustomDigitModel()

# ================================================================
#  Okuma Mantığı (Sadece Özel Model)
# ================================================================
def read_display(frame, config):
    if not custom_model.is_ready():
        return "", 0.0, {}

    x, y, w, h = config["x"], config["y"], config["w"], config["h"]
    n_digits    = config.get("n_digits", 5)
    roi = frame[y:y+h, x:x+w]
    step = w // n_digits

    custom_text = ""
    custom_conf = 0.0
    digit_chars = []

    for i in range(n_digits):
        # Ayarlanmış slot sınırlarını okur
        if "slots" in config and i < len(config["slots"]):
            sx = config["slots"][i]["x"]
            sw = config["slots"][i]["w"]
        else:
            sx = i * step
            sw = step
            
        digit_roi = roi[:, sx : sx + sw]
        d32       = preprocess_digit(digit_roi)
        label, conf = custom_model.predict(d32)
        digit_chars.append((label, conf))

    # "bos" etiketli bölmeleri atla, sadece rakamları birleştir
    custom_text = "".join(str(lbl) for lbl, _ in digit_chars if lbl != "bos")
    
    display_text = ""
    float_value = None
    
    if custom_text:
        custom_conf = float(np.mean([c for l, c in digit_chars if l != "bos"]))
        try:
            raw_number = float(custom_text)
            
            # BURADAKİ 10.0 DEĞERİNİ İHTİYACINIZA GÖRE 100.0 YAPABİLİRSİNİZ
            float_value = raw_number / 100.0 
            
            display_text = str(float_value)
        except ValueError:
            display_text = custom_text

    debug_info = {
        "raw_text": custom_text,
        "custom_conf": custom_conf,
        "float_value": float_value,
    }
    
    return display_text, custom_conf, debug_info

# ================================================================
#  Excel kayıt — Görünen değeri kaydeder
# ================================================================
def save_to_excel(display_text, confidence=None, excel_file=EXCEL_FILE):
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LCD Kayitlari"

    try:
        val_to_save = float(display_text)
    except ValueError:
        val_to_save = display_text
        
    ws.append([val_to_save])
    wb.save(excel_file)
    return ws.max_row - 1

# ================================================================
#  Kalibrasyon
# ================================================================
def calibrate(cap):
    print("\n[KALIBRASYON] Fare ile ekrani secin, ENTER=onayla, ESC=iptal")
    for _ in range(5): cap.read()
    ret, frame = cap.read()
    if not ret: return None

    roi_coords = cv2.selectROI(
        "Ekrani Secin: Surukle, ENTER=Onayla, ESC=Iptal",
        frame, fromCenter=False, showCrosshair=True
    )
    cv2.destroyAllWindows()

    x, y, w, h = [int(v) for v in roi_coords]
    if w == 0 or h == 0: return None

    print("\n[ONEMLI]: Ekrana sigabilecek MAKSIMUM rakam sayisini girin.")
    n_str = input("Maksimum rakam sayisi kactir? (Orn: 5): ").strip()
    n_digits = int(n_str) if n_str.isdigit() else 5

    config = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            config = json.load(f)
    config.update({"x": x, "y": y, "w": w, "h": h, "n_digits": n_digits})
    
    # Yeni slotları sıfırla ki manuel toplayıcıda düzgün başlasın
    step = w // n_digits
    config["slots"] = [{"x": i * step, "w": step} for i in range(n_digits)]
    
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)
    print(f"[OK] Kaydedildi -> {CONFIG_FILE}")
    return config

# ================================================================
#  Ana program
# ================================================================
def main():
    print()
    print("=" * 60)
    print("  LCD Okuyucu v5  --  EasyOCR'siz Hizli Model (/10 bolmeli)")
    print("  [S] Kaydet  [A] Oto  [C] Kalibrasyon  [Q] Cik")
    print("=" * 60)

    cap = cv2.VideoCapture(1)
    if not cap.isOpened():
        print("[HATA] Kamera acilamadi!")
        return
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

    config = None
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            config = json.load(f)

    if not config or not all(k in config for k in ("x", "y", "w", "h")):
        config = calibrate(cap)
        if config is None:
            cap.release()
            return

    auto_save     = False
    auto_interval = 5
    last_auto_t   = 0
    last_reading  = ""
    last_conf     = 0.0
    save_count    = 0
    frame_skip    = 0
    OCR_EVERY     = 2  # EasyOCR olmadığı için daha sık okuma yapabiliriz
    last_debug    = {}

    while True:
        ret, frame = cap.read()
        if not ret: break

        display = frame.copy()
        x, y, w, h = config["x"], config["y"], config["w"], config["h"]

        frame_skip += 1
        if frame_skip >= OCR_EVERY:
            frame_skip = 0
            reading, conf, debug = read_display(frame, config)
            if reading:
                last_reading = reading
                last_conf    = conf
                last_debug   = debug

        # ROI kutusu
        color = (0, 220, 80) if last_reading else (80, 80, 80)
        cv2.rectangle(display, (x, y), (x+w, y+h), color, 2)

        # Ayarlanmış slot çizgilerini göstermek isterseniz (opsiyonel)
        if "slots" in config:
            for i, slot in enumerate(config["slots"]):
                sx = x + slot["x"]
                sw = slot["w"]
                cv2.rectangle(display, (sx, y), (sx+sw, y+h), (200, 100, 0), 1)

        if last_reading:
            conf_pct  = int(last_conf * 100)
            label     = f"{last_reading}  ({conf_pct}%)"

            (lw, lh), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.9, 2)
            cv2.rectangle(display, (x, y - lh - 16), (x + lw + 12, y), (0, 0, 0), -1)
            cv2.putText(display, label, (x+6, y-8), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)

            # Debug satırı
            raw_t = last_debug.get("raw_text", "")
            fval  = last_debug.get("float_value", None)
            if raw_t:
                debug_line = f"Ham Okunan: {raw_t} -> Bolunmus Deger: {fval}"
                cv2.putText(display, debug_line, (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (180, 180, 80), 2)

        if auto_save and last_reading:
            now_t = time.time()
            if now_t - last_auto_t >= auto_interval:
                save_count  = save_to_excel(last_reading, last_conf)
                last_auto_t = now_t
                print(f"[OTO] {last_reading}  {int(last_conf*100)}%  satir={save_count}")

        H, W = display.shape[:2]
        cv2.rectangle(display, (0, H-90), (W, H), (20, 20, 20), -1)
        auto_str = f"ACIK({auto_interval}s)" if auto_save else "KAPALI"
        cv2.putText(display, f"[S]Kaydet [A]Oto:{auto_str} [+/-]Aralik [C]Kalibrasyon [Q]Cik",
                    (10, H-58), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (180, 180, 180), 1)
        model_status = "Ozel model: AKTIF" if custom_model.is_ready() else "Ozel model: YOK"
        cv2.putText(display, f"Son: {last_reading or '-'} | Kayit:{save_count} | {model_status}",
                    (10, H-28), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (100, 230, 130), 1)

        cv2.imshow("LCD Okuyucu v5", display)
        key = cv2.waitKey(1) & 0xFF

        if key in (ord("q"), ord("Q")): break
        elif key in (ord("s"), ord("S")):
            if last_reading:
                save_count = save_to_excel(last_reading, last_conf)
                print(f"[KAYIT] {last_reading}  {int(last_conf*100)}%  satir={save_count}")
        elif key in (ord("a"), ord("A")):
            auto_save = not auto_save
        elif key == ord("+"): auto_interval = min(auto_interval + 1, 60)
        elif key == ord("-"): auto_interval = max(auto_interval - 1, 1)
        elif key in (ord("c"), ord("C")): config = calibrate(cap)

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()